import re
import os
import time
import requests
import feedparser

from datetime import datetime
from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright

# =====================================
# CONFIG
# =====================================

RSS_URL = "https://www.securityweek.com/category/vulnerabilities/feed/"

FILE = "securityweek_cves.xlsx"

CVE_PATTERN = r"CVE[\s\-–—]\d{4}[\s\-–—]\d{4,7}"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 "
        "(Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 "
        "(KHTML, like Gecko) "
        "Chrome/136 Safari/537.36"
    )
}

# =====================================
# CLEAN CVE
# =====================================

def clean_cve(cve):

    cve = cve.upper()

    cve = re.sub(r"[\s_–—]+", "-", cve)

    cve = re.sub(r"-+", "-", cve)

    return cve.strip("-")

# =====================================
# EXTRACT CVEs
# =====================================

def extract_cves(playwright, url):

    browser = playwright.chromium.launch(headless=True)

    page = browser.new_page()

    try:

        page.goto(
            url,
            wait_until="domcontentloaded",
            timeout=60000
        )

        page.wait_for_timeout(4000)

        html = page.content()

    except Exception as e:

        print("FAILED:", url)
        print(e)

        browser.close()

        return set()

    browser.close()

    raw = re.findall(
        CVE_PATTERN,
        html,
        re.IGNORECASE
    )

    return {
        clean_cve(c)
        for c in raw
    }

# =====================================
# LOAD EXISTING ROWS
# =====================================

def load_existing_rows():

    if not os.path.exists(FILE):
        return []

    wb = load_workbook(FILE)

    ws = wb.active

    rows = []

    for row in ws.iter_rows(
        min_row=2,
        values_only=True
    ):

        if row[0]:
            rows.append(row)

    return rows

# =====================================
# EXISTING UNIQUE KEYS
# =====================================

def load_existing_keys():

    rows = load_existing_rows()

    return {
        (r[0], r[2])
        for r in rows
    }

# =====================================
# SAVE ALL
# =====================================

def save_all(rows):

    wb = Workbook()

    ws = wb.active

    ws.append([
        "CVE",
        "DATE",
        "TITLE",
        "LINK"
    ])

    # oldest -> newest
    rows_sorted = sorted(
        rows,
        key=lambda x: datetime.strptime(
            x[1],
            "%Y-%m-%d"
        )
    )

    for r in rows_sorted:
        ws.append(r)

    wb.save(FILE)

# =====================================
# MAIN
# =====================================

def main():

    print("\nFetching RSS feed...\n")

    response = requests.get(
        RSS_URL,
        headers=HEADERS,
        timeout=30
    )

    feed = feedparser.parse(
        response.content
    )

    print("Articles Found:", len(feed.entries))

    existing_rows = load_existing_rows()

    existing_keys = load_existing_keys()

    new_rows = []

    with sync_playwright() as playwright:

        for idx, entry in enumerate(feed.entries, start=1):

            if not hasattr(
                entry,
                "published_parsed"
            ):
                continue

            pub = datetime(
                *entry.published_parsed[:6]
            )

            date_str = pub.strftime(
                "%Y-%m-%d"
            )

            title = entry.title.strip()

            link = entry.link.strip()

            print("\n" + "=" * 70)

            print(f"ARTICLE #{idx}")

            print("=" * 70)

            print("DATE :", date_str)

            print("TITLE:", title)

            print("LINK :", link)

            cves = extract_cves(
                playwright,
                link
            )

            if cves:

                print("\nCVEs Found:")

                for c in cves:

                    print("-", c)

                    key = (c, link)

                    if key not in existing_keys:

                        new_rows.append((
                            c,
                            date_str,
                            title,
                            link
                        ))

                        existing_keys.add(key)

            else:

                print("No CVEs Found")

            time.sleep(2)

    all_rows = existing_rows + new_rows

    if new_rows:

        save_all(all_rows)

        print(
            f"\nAdded {len(new_rows)} new CVEs"
        )

    else:

        print("\nNo new CVEs")

# =====================================
# RUN
# =====================================

if _name_ == "_main_":
    main()
